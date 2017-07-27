#!/usr/bin/env python  
# encoding: utf-8  
"""
@author: Alfons
@contact: alfons_xh@163.com
@file: WriteXlsx.py 
@time: 2017/7/26 16:33 
@version: v1.0 
"""
from openpyxl import Workbook
import xlrd
import xlwt
import time


class Write_Openxls:
    def __init__(self):
        pass

    def Write_xls(self, xls_name):
        wb = Workbook(write_only = True)
        for j in range(2 ** 4):
            ws = wb.create_sheet("Test_sheet" + str(j))
            for i in range(1, 2 ** 16):
                ws.append(["hello", "hello", "hello", "hello"])
        wb.save(xls_name)
        pass


class Write_xlwt:
    def __init__(self):
        pass

    def Write_xls(self, xls_name):
        wb = xlwt.Workbook()
        for sheet_number in range(2 ** 4):
            ws = wb.add_sheet("Test_sheet" + str(sheet_number), cell_overwrite_ok = True)
            for i in range(2 ** 16):
                ws.write(i, 0, "hello")
                ws.write(i, 2, "hello")
                ws.write(i, 1, "hello")
                ws.write(i, 3, "hello")
        wb.save(xls_name)
        pass


if __name__ == "__main__":
    start_time = time.clock()
    openxls = Write_Openxls()
    openxls.Write_xls("openpyxl_test.xls")
    end_time = time.clock()
    print "use: %f s" % (end_time - start_time)

    start_time = time.clock()
    xl_wt = Write_xlwt()
    xl_wt.Write_xls("xlwt_test.xls")
    end_time = time.clock()
    print "use: %f s" % (end_time - start_time)
    pass
